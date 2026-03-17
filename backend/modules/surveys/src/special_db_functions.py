# Special Survey Module – Database Functions
# Handles the 3-phase workflow for new class composition surveys.

import csv
import io
from datetime import datetime, timezone
from html import escape as html_escape

from src.db import db
from src.db_models import User
from modules.surveys.src.db_models import (
    Survey, SurveyResponse, SurveyAnswer,
    SpecialSurvey, SpecialSurveyStudent, SpecialSurveyParent,
    SpecialSurveyClassTeacher, SpecialSurveyStudentWish,
    SpecialSurveyTeacherEvaluation,
)

# Maximum rows allowed in a single CSV upload
_CSV_MAX_ROWS = 2000


def _get_survey_for_owner(special_survey_id, creator_uuid):
    """
    Helper: fetch a SpecialSurvey and verify ownership.
    Returns (survey, None) on success or (None, error_dict) on failure.
    """
    ss = SpecialSurvey.query.get(special_survey_id)
    if not ss:
        return None, {'status': False, 'message': 'Umfrage nicht gefunden.'}
    if ss.creator_uuid != creator_uuid:
        return None, {'status': False, 'message': 'Keine Berechtigung.'}
    return ss, None





def migrate_class_teacher_constraint():
    """
    Migrate the special_survey_class_teacher table to allow multiple
    teachers per class (unique on survey_id + class_name + teacher_uuid
    instead of survey_id + class_name).
    Only needed for SQLite where ALTER TABLE cannot drop constraints.
    """
    from sqlalchemy import inspect, text
    try:
        inspector = inspect(db.engine)
        if 'special_survey_class_teacher' not in inspector.get_table_names():
            return  # Table doesn't exist yet, create_all will handle it

        # Check current unique constraints
        ucs = inspector.get_unique_constraints('special_survey_class_teacher')
        needs_migration = False
        for uc in ucs:
            if uc.get('name') == 'uq_special_class_teacher':
                cols = uc.get('column_names', [])
                if 'teacher_uuid' not in cols:
                    needs_migration = True
                break

        if not needs_migration:
            return

        # SQLite: recreate table with new constraint
        with db.engine.begin() as conn:
            # 1. Save existing data
            rows = conn.execute(text(
                'SELECT id, special_survey_id, class_name, teacher_uuid '
                'FROM special_survey_class_teacher'
            )).fetchall()

            # 2. Drop old table
            conn.execute(text('DROP TABLE special_survey_class_teacher'))

            # 3. Recreate with new constraint (create_all will do this)
            # but we need the table now
            conn.execute(text('''
                CREATE TABLE special_survey_class_teacher (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    special_survey_id INTEGER NOT NULL,
                    class_name VARCHAR(100) NOT NULL,
                    teacher_uuid VARCHAR NOT NULL,
                    FOREIGN KEY (special_survey_id) REFERENCES special_survey(id) ON DELETE CASCADE,
                    FOREIGN KEY (teacher_uuid) REFERENCES user(uuid),
                    UNIQUE (special_survey_id, class_name, teacher_uuid)
                )
            '''))

            # 4. Re-insert data
            for row in rows:
                conn.execute(text(
                    'INSERT INTO special_survey_class_teacher '
                    '(id, special_survey_id, class_name, teacher_uuid) '
                    'VALUES (:id, :ssid, :cn, :tu)'
                ), {'id': row[0], 'ssid': row[1], 'cn': row[2], 'tu': row[3]})

        print("[Special Survey] Migrated class_teacher constraint for multi-teacher support.")
    except Exception as e:
        print(f"[Special Survey] Migration check skipped: {e}")


def migrate_template_type_and_excel_config():
    """
    Add template_type column to survey table and excel_config_json to
    survey_question table if they don't already exist.
    Needed for the teacher evaluation template separation feature.
    """
    from sqlalchemy import inspect, text
    try:
        inspector = inspect(db.engine)

        # 1. Add template_type to survey table
        if 'survey' in inspector.get_table_names():
            columns = [c['name'] for c in inspector.get_columns('survey')]
            if 'template_type' not in columns:
                with db.engine.begin() as conn:
                    conn.execute(text(
                        "ALTER TABLE survey ADD COLUMN template_type VARCHAR(30) NOT NULL DEFAULT 'normal'"
                    ))
                print("[Surveys] Added template_type column to survey table.")

        # 2. Add excel_config_json to survey_question table
        if 'survey_question' in inspector.get_table_names():
            columns = [c['name'] for c in inspector.get_columns('survey_question')]
            if 'excel_config_json' not in columns:
                with db.engine.begin() as conn:
                    conn.execute(text(
                        "ALTER TABLE survey_question ADD COLUMN excel_config_json TEXT DEFAULT '{}'"
                    ))
                print("[Surveys] Added excel_config_json column to survey_question table.")

    except Exception as e:
        print(f"[Surveys] Template type migration skipped: {e}")


# ── CSV Required Columns ────────────────────────────────────────────

REQUIRED_CSV_COLUMNS = {'Account', 'Vorname', 'Nachname'}
STUDENT_EXTRA_COLUMNS = {'Klasse/Information'}


def _parse_csv(file_content, is_student=False):
    """
    Parse a CSV file with semicolon delimiter.
    Returns (list_of_dicts, error_string_or_None).
    """
    try:
        text = file_content.decode('utf-8-sig')  # handle BOM
    except UnicodeDecodeError:
        try:
            text = file_content.decode('latin-1')
        except UnicodeDecodeError:
            return None, 'CSV-Datei konnte nicht gelesen werden (ungültige Zeichenkodierung).'

    reader = csv.DictReader(io.StringIO(text), delimiter=';')
    if not reader.fieldnames:
        return None, 'CSV-Datei ist leer oder hat keine Kopfzeile.'

    # Strip surrounding quotes/whitespace from fieldnames
    reader.fieldnames = [f.strip().strip('"') for f in reader.fieldnames]

    required = REQUIRED_CSV_COLUMNS.copy()
    if is_student:
        required |= STUDENT_EXTRA_COLUMNS

    missing = required - set(reader.fieldnames)
    if missing:
        return None, f'Fehlende Spalten: {", ".join(sorted(missing))}'

    rows = []
    for i, row in enumerate(reader, start=2):
        # Strip whitespace and quotes from values
        cleaned = {k.strip().strip('"'): (v.strip().strip('"') if v else '') for k, v in row.items() if k}
        account = cleaned.get('Account', '').strip()
        if not account:
            continue  # skip empty rows

        entry = {
            'account': account,
            'first_name': cleaned.get('Vorname', ''),
            'last_name': cleaned.get('Nachname', ''),
            'email': cleaned.get('E-Mail-Adresse', ''),
        }
        if is_student:
            class_info = cleaned.get('Klasse/Information', '').strip()
            if not class_info:
                return None, f'Zeile {i}: "Klasse/Information" fehlt für Account "{account}".'
            entry['class_name'] = class_info

        rows.append(entry)

    if not rows:
        return None, 'CSV-Datei enthält keine gültigen Einträge.'

    if len(rows) > _CSV_MAX_ROWS:
        return None, f'CSV-Datei enthält zu viele Einträge (max. {_CSV_MAX_ROWS}).'

    return rows, None


# ── Create Special Survey (Phase 0) ────────────────────────────────

def create_special_survey(title, description, creator_uuid, grade_level,
                          student_csv_content, parent_csv_content,
                          teacher_survey_id=None):
    """
    Create a new special survey by uploading student and parent CSVs.
    Returns dict with status, message, and survey data.
    """
    try:
        creator = User.query.filter_by(uuid=creator_uuid).first()
        if not creator:
            return {'status': False, 'message': 'Ersteller nicht gefunden.'}

        # Parse CSVs
        students_data, err = _parse_csv(student_csv_content, is_student=True)
        if err:
            return {'status': False, 'message': f'Schüler-CSV: {err}'}

        parents_data, err = _parse_csv(parent_csv_content, is_student=False)
        if err:
            return {'status': False, 'message': f'Eltern-CSV: {err}'}

        # Create special survey
        ss = SpecialSurvey(
            title=title,
            description=description or '',
            creator_uuid=creator_uuid,
            grade_level=grade_level,
            current_phase=0,
            status='setup',
            teacher_survey_id=teacher_survey_id,
        )
        db.session.add(ss)
        db.session.flush()

        # Import students
        for s_data in students_data:
            user = User.query.filter_by(username=s_data['account']).first()
            student = SpecialSurveyStudent(
                special_survey_id=ss.id,
                account=s_data['account'],
                first_name=s_data['first_name'],
                last_name=s_data['last_name'],
                class_name=s_data['class_name'],
                user_uuid=user.uuid if user else None,
            )
            db.session.add(student)

        # Import parents
        for p_data in parents_data:
            user = User.query.filter_by(username=p_data['account']).first()
            parent = SpecialSurveyParent(
                special_survey_id=ss.id,
                account=p_data['account'],
                first_name=p_data['first_name'],
                last_name=p_data['last_name'],
                email=p_data.get('email', ''),
                user_uuid=user.uuid if user else None,
            )
            db.session.add(parent)

        db.session.commit()

        return {
            'status': True,
            'message': 'Spezialumfrage erstellt.',
            'survey_id': ss.id,
            'student_count': len(students_data),
            'parent_count': len(parents_data),
        }

    except Exception as e:
        db.session.rollback()
        print(f'[Surveys] Error: {e}')
        return {'status': False, 'message': 'Fehler beim Erstellen.'}


def get_special_survey_classes(special_survey_id):
    """Return list of unique class names and their student counts."""
    ss = SpecialSurvey.query.get(special_survey_id)
    if not ss:
        return None

    classes = {}
    for s in ss.students:
        if s.class_name not in classes:
            classes[s.class_name] = {'class_name': s.class_name, 'student_count': 0, 'students': []}
        classes[s.class_name]['student_count'] += 1
        classes[s.class_name]['students'].append(s.to_dict())

    # Include existing teacher assignments (multiple per class)
    teacher_map = {}  # class_name -> list of teacher dicts
    for ct in ss.class_teachers:
        teacher_map.setdefault(ct.class_name, []).append({
            'teacher_uuid': ct.teacher_uuid,
            'teacher_name': ct.teacher.username if ct.teacher else None,
        })

    result = []
    for cn in sorted(classes.keys()):
        c = classes[cn]
        c['teachers'] = teacher_map.get(cn, [])
        # Backward compat: keep single 'teacher' as first assigned (or None)
        c['teacher'] = teacher_map.get(cn, [None])[0] if teacher_map.get(cn) else None
        result.append(c)

    return result


def assign_class_teachers(special_survey_id, assignments, creator_uuid):
    """
    Assign teachers to classes.
    `assignments` is a list of { class_name, teacher_uuids: [...] }
    or legacy format { class_name, teacher_uuid }.
    Multiple teachers per class are supported.
    """
    try:
        ss, err = _get_survey_for_owner(special_survey_id, creator_uuid)
        if err:
            return err

        # Clear existing assignments
        SpecialSurveyClassTeacher.query.filter_by(special_survey_id=special_survey_id).delete()

        for a in assignments:
            class_name = a.get('class_name', '').strip()
            if not class_name:
                continue

            # Support both formats: teacher_uuids (list) and teacher_uuid (single)
            teacher_uuids = a.get('teacher_uuids', [])
            if not teacher_uuids:
                single = a.get('teacher_uuid', '').strip()
                if single:
                    teacher_uuids = [single]

            for t_uuid in teacher_uuids:
                t_uuid = t_uuid.strip() if isinstance(t_uuid, str) else ''
                if not t_uuid:
                    continue

                # Validate teacher exists
                teacher = User.query.filter_by(uuid=t_uuid).first()
                if not teacher:
                    return {'status': False, 'message': f'Lehrkraft nicht gefunden: {t_uuid}'}

                ct = SpecialSurveyClassTeacher(
                    special_survey_id=special_survey_id,
                    class_name=class_name,
                    teacher_uuid=t_uuid,
                )
                db.session.add(ct)

        db.session.commit()
        return {'status': True, 'message': 'Klassenlehrkräfte zugewiesen.'}

    except Exception as e:
        db.session.rollback()
        print(f'[Surveys] Error: {e}')
        return {'status': False, 'message': 'Ein Fehler ist aufgetreten.'}


def activate_survey(special_survey_id, creator_uuid):
    """
    Activate the special survey: transition from setup → active.
    """
    try:
        ss, err = _get_survey_for_owner(special_survey_id, creator_uuid)
        if err:
            return err

        if ss.status != 'setup':
            return {'status': False, 'message': 'Umfrage ist bereits aktiviert.'}

        # Must have class teachers assigned
        if not ss.class_teachers:
            return {'status': False, 'message': 'Bitte weisen Sie zuerst Klassenlehrkräfte zu.'}

        ss.status = 'active'
        ss.current_phase = 1  # keep for backward compat, but 'active' status is the gate

        db.session.commit()
        return {
            'status': True,
            'message': 'Umfrage aktiviert – alle Rollen können nun teilnehmen.',
            'new_status': 'active',
        }

    except Exception as e:
        db.session.rollback()
        print(f'[Surveys] Error: {e}')
        return {'status': False, 'message': 'Ein Fehler ist aufgetreten.'}


def complete_survey(special_survey_id, creator_uuid):
    """
    Complete the special survey: transition from active → completed.
    No further participation possible after this.
    """
    try:
        ss, err = _get_survey_for_owner(special_survey_id, creator_uuid)
        if err:
            return err

        if ss.status != 'active':
            return {'status': False, 'message': 'Umfrage ist nicht aktiv.'}

        ss.status = 'completed'

        db.session.commit()
        return {
            'status': True,
            'message': 'Umfrage abgeschlossen.',
            'new_status': 'completed',
        }

    except Exception as e:
        db.session.rollback()
        print(f'[Surveys] Error: {e}')
        return {'status': False, 'message': 'Ein Fehler ist aufgetreten.'}


def archive_special_survey(special_survey_id, creator_uuid):
    """
    Archive a special survey.
    Allowed from 'completed' or 'active' status.
    Preserves all data; export remains functional.
    """
    try:
        ss, err = _get_survey_for_owner(special_survey_id, creator_uuid)
        if err:
            return err

        if ss.status not in ('completed', 'active'):
            return {'status': False, 'message': 'Nur aktive oder abgeschlossene Umfragen können archiviert werden.'}

        ss.status = 'archived'
        db.session.commit()
        return {
            'status': True,
            'message': 'Umfrage archiviert.',
            'new_status': 'archived',
        }

    except Exception as e:
        db.session.rollback()
        print(f'[Surveys] Error: {e}')
        return {'status': False, 'message': 'Ein Fehler ist aufgetreten.'}


def reactivate_special_survey(special_survey_id, creator_uuid):
    """
    Reactivate a completed or archived special survey back to 'active'.
    All existing data (wishes, evaluations, etc.) is preserved.
    """
    try:
        ss, err = _get_survey_for_owner(special_survey_id, creator_uuid)
        if err:
            return err

        if ss.status not in ('completed', 'archived'):
            return {'status': False, 'message': 'Nur abgeschlossene oder archivierte Umfragen können reaktiviert werden.'}

        ss.status = 'active'
        db.session.commit()
        return {
            'status': True,
            'message': 'Umfrage reaktiviert.',
            'new_status': 'active',
        }

    except Exception as e:
        db.session.rollback()
        print(f'[Surveys] Error: {e}')
        return {'status': False, 'message': 'Ein Fehler ist aufgetreten.'}


# Legacy alias for backward compatibility
def advance_phase(special_survey_id, creator_uuid):
    """Legacy wrapper: activates or completes the survey."""
    ss = SpecialSurvey.query.get(special_survey_id)
    if not ss:
        return {'status': False, 'message': 'Umfrage nicht gefunden.'}
    if ss.status == 'setup':
        return activate_survey(special_survey_id, creator_uuid)
    elif ss.status == 'active':
        return complete_survey(special_survey_id, creator_uuid)
    else:
        return {'status': False, 'message': 'Ungültiger Status für Phasenwechsel.'}


# ── Phase 1: Student Input ──────────────────────────────────────────

def get_student_phase1_data(special_survey_id, user_uuid):
    """
    Get Phase 1 data for a student: their classmates to choose from,
    parent list, and any existing wishes.
    """
    ss = SpecialSurvey.query.get(special_survey_id)
    if not ss:
        return None, 'Umfrage nicht gefunden.'
    if ss.status not in ('active',):
        return None, 'Die Umfrage ist nicht aktiv.'

    # Lazy-link in case user_uuid was NULL at import time
    user = User.query.filter_by(uuid=user_uuid).first()
    if user:
        _lazy_link_user(user)

    # Find the student record
    student = SpecialSurveyStudent.query.filter_by(
        special_survey_id=special_survey_id, user_uuid=user_uuid
    ).first()
    if not student:
        return None, 'Sie sind nicht als Schüler/in in dieser Umfrage registriert.'

    # Check if wishes are locked (parent confirmed)
    existing_wish = SpecialSurveyStudentWish.query.filter_by(
        special_survey_id=special_survey_id, student_id=student.id
    ).first()
    if existing_wish and existing_wish.locked:
        return None, 'Ihre Wünsche wurden bereits von einem Elternteil bestätigt und können nicht mehr geändert werden.'

    # All students in the same grade (not just same class)
    all_students = SpecialSurveyStudent.query.filter(
        SpecialSurveyStudent.special_survey_id == special_survey_id,
        SpecialSurveyStudent.id != student.id,  # exclude self
    ).order_by(SpecialSurveyStudent.last_name, SpecialSurveyStudent.first_name).all()

    # All parents
    all_parents = SpecialSurveyParent.query.filter_by(
        special_survey_id=special_survey_id
    ).order_by(SpecialSurveyParent.last_name, SpecialSurveyParent.first_name).all()

    return {
        'student': student.to_dict(),
        'classmates': [s.to_dict() for s in all_students],
        'parents': [p.to_dict() for p in all_parents],
        'existing_wish': existing_wish.to_dict() if existing_wish else None,
        'survey_title': ss.title,
    }, None


def submit_student_wishes(special_survey_id, user_uuid, wish1_student_id, wish2_student_id, selected_parent_id):
    """Submit or update a student's wishes in Phase 1."""
    try:
        # Validate ID types
        try:
            wish1_student_id = int(wish1_student_id)
            wish2_student_id = int(wish2_student_id)
            selected_parent_id = int(selected_parent_id)
        except (TypeError, ValueError):
            return {'status': False, 'message': 'Ungültige Eingabedaten.'}

        ss = SpecialSurvey.query.get(special_survey_id)
        if not ss:
            return {'status': False, 'message': 'Umfrage nicht gefunden.'}
        if ss.status not in ('active',):
            return {'status': False, 'message': 'Die Umfrage ist nicht aktiv.'}

        # Find student
        student = SpecialSurveyStudent.query.filter_by(
            special_survey_id=special_survey_id, user_uuid=user_uuid
        ).first()
        if not student:
            return {'status': False, 'message': 'Sie sind nicht als Schüler/in registriert.'}

        # Check not locked
        existing = SpecialSurveyStudentWish.query.filter_by(
            special_survey_id=special_survey_id, student_id=student.id
        ).first()
        if existing and existing.locked:
            return {'status': False, 'message': 'Ihre Wünsche sind gesperrt (Eltern haben bestätigt).'}

        # Validate wishes
        if wish1_student_id == wish2_student_id:
            return {'status': False, 'message': 'Die beiden Wünsche müssen unterschiedlich sein.'}
        if wish1_student_id == student.id or wish2_student_id == student.id:
            return {'status': False, 'message': 'Sie können sich nicht selbst auswählen.'}

        # Validate wish students exist in this survey
        w1 = SpecialSurveyStudent.query.filter_by(
            id=wish1_student_id, special_survey_id=special_survey_id
        ).first()
        w2 = SpecialSurveyStudent.query.filter_by(
            id=wish2_student_id, special_survey_id=special_survey_id
        ).first()
        if not w1 or not w2:
            return {'status': False, 'message': 'Ungültige Schülerauswahl.'}

        # Validate parent exists in this survey
        parent = SpecialSurveyParent.query.filter_by(
            id=selected_parent_id, special_survey_id=special_survey_id
        ).first()
        if not parent:
            return {'status': False, 'message': 'Ungültiger Elternaccount.'}

        if existing:
            existing.wish1_student_id = wish1_student_id
            existing.wish2_student_id = wish2_student_id
            existing.selected_parent_id = selected_parent_id
            existing.parent_confirmed = False  # reset confirmation on edit
        else:
            wish = SpecialSurveyStudentWish(
                special_survey_id=special_survey_id,
                student_id=student.id,
                wish1_student_id=wish1_student_id,
                wish2_student_id=wish2_student_id,
                selected_parent_id=selected_parent_id,
                parent_confirmed=False,
                locked=False,
            )
            db.session.add(wish)

        db.session.commit()

        # Send email notification to the selected parent (non-blocking, non-fatal)
        student_name = f"{student.first_name} {student.last_name}"
        send_parent_notification_email(selected_parent_id, special_survey_id, student_name)

        return {'status': True, 'message': 'Wünsche gespeichert.'}

    except Exception as e:
        db.session.rollback()
        print(f'[Surveys] Error: {e}')
        return {'status': False, 'message': 'Ein Fehler ist aufgetreten.'}


# ── Phase 2: Parent Confirmation ────────────────────────────────────

def get_parent_phase2_data(special_survey_id, user_uuid):
    """
    Get Phase 2 data for a parent: the child's wishes to confirm.
    Only accessible if the parent was selected in Phase 1.
    """
    ss = SpecialSurvey.query.get(special_survey_id)
    if not ss:
        return None, 'Umfrage nicht gefunden.'
    if ss.status not in ('active',):
        return None, 'Die Umfrage ist nicht aktiv.'

    # Lazy-link in case user_uuid was NULL at import time
    user = User.query.filter_by(uuid=user_uuid).first()
    if user:
        _lazy_link_user(user)

    # Find parent record by user_uuid
    parent = SpecialSurveyParent.query.filter_by(
        special_survey_id=special_survey_id, user_uuid=user_uuid
    ).first()
    if not parent:
        return None, 'Sie sind nicht als Elternteil in dieser Umfrage registriert.'

    # Find wishes where this parent was selected
    wishes = SpecialSurveyStudentWish.query.filter_by(
        special_survey_id=special_survey_id,
        selected_parent_id=parent.id,
    ).all()

    if not wishes:
        return None, 'Sie wurden von keinem Kind ausgewählt.'

    children_data = []
    for wish in wishes:
        children_data.append({
            'wish_id': wish.id,
            'student': wish.student.to_dict(),
            'wish1': wish.wish1_student.to_dict() if wish.wish1_student else None,
            'wish2': wish.wish2_student.to_dict() if wish.wish2_student else None,
            'parent_confirmed': wish.parent_confirmed,
            'locked': wish.locked,
        })

    return {
        'parent': parent.to_dict(),
        'children': children_data,
        'survey_title': ss.title,
    }, None


def confirm_parent_wishes(special_survey_id, user_uuid, wish_id):
    """Parent confirms a child's wishes. Locks the wishes."""
    try:
        # Validate ID type
        try:
            wish_id = int(wish_id)
        except (TypeError, ValueError):
            return {'status': False, 'message': 'Ungültige Wunsch-ID.'}

        ss = SpecialSurvey.query.get(special_survey_id)
        if not ss:
            return {'status': False, 'message': 'Umfrage nicht gefunden.'}
        if ss.status not in ('active',):
            return {'status': False, 'message': 'Die Umfrage ist nicht aktiv.'}

        parent = SpecialSurveyParent.query.filter_by(
            special_survey_id=special_survey_id, user_uuid=user_uuid
        ).first()
        if not parent:
            return {'status': False, 'message': 'Keine Berechtigung.'}

        wish = SpecialSurveyStudentWish.query.get(wish_id)
        if not wish or wish.special_survey_id != special_survey_id:
            return {'status': False, 'message': 'Wunsch nicht gefunden.'}
        if wish.selected_parent_id != parent.id:
            return {'status': False, 'message': 'Keine Berechtigung für diesen Wunsch.'}

        wish.parent_confirmed = True
        wish.locked = True
        db.session.commit()
        return {'status': True, 'message': 'Wünsche bestätigt und gesperrt.'}

    except Exception as e:
        db.session.rollback()
        print(f'[Surveys] Error: {e}')
        return {'status': False, 'message': 'Ein Fehler ist aufgetreten.'}


# ── Phase 3: Teacher Evaluation ─────────────────────────────────────

def get_teacher_phase3_data(special_survey_id, user_uuid):
    """
    Get Phase 3 data for a teacher: students in their assigned class and
    the teacher survey template questions.
    Teachers answer ONLY the questions defined in the linked survey template.
    """
    ss = SpecialSurvey.query.get(special_survey_id)
    if not ss:
        return None, 'Umfrage nicht gefunden.'
    if ss.status not in ('active',):
        return None, 'Die Umfrage ist nicht aktiv.'

    # Find class assignment for this teacher
    assignments = SpecialSurveyClassTeacher.query.filter_by(
        special_survey_id=special_survey_id, teacher_uuid=user_uuid
    ).all()
    if not assignments:
        return None, 'Sie sind keiner Klasse in dieser Umfrage zugewiesen.'

    classes_data = []
    for assignment in assignments:
        students = SpecialSurveyStudent.query.filter_by(
            special_survey_id=special_survey_id,
            class_name=assignment.class_name,
        ).order_by(SpecialSurveyStudent.last_name, SpecialSurveyStudent.first_name).all()

        # Get existing evaluations with their survey response answers
        existing_evals = {}
        for s in students:
            ev = SpecialSurveyTeacherEvaluation.query.filter_by(
                special_survey_id=special_survey_id,
                student_id=s.id,
            ).first()
            if ev:
                ev_data = ev.to_dict()
                # Include existing answers from linked survey response
                if ev.survey_response_id and ev.survey_response:
                    answers_map = {}
                    for a in ev.survey_response.answers:
                        answers_map[a.question_id] = a.to_dict()
                    ev_data['answers'] = answers_map
                existing_evals[s.id] = ev_data

        classes_data.append({
            'class_name': assignment.class_name,
            'students': [s.to_dict() for s in students],
            'existing_evaluations': existing_evals,
        })

    # Get teacher survey template questions (the ONLY questions teachers answer)
    teacher_questions = []
    if ss.teacher_survey_id:
        survey = Survey.query.get(ss.teacher_survey_id)
        if survey:
            teacher_questions = [q.to_dict() for q in survey.questions]

    return {
        'classes': classes_data,
        'teacher_questions': teacher_questions,
        'survey_title': ss.title,
    }, None


def submit_teacher_evaluation(special_survey_id, user_uuid, student_id,
                              survey_answers=None, **_legacy_kwargs):
    """
    Submit or update a teacher's evaluation of a student.
    Teachers answer ONLY the questions from the linked survey template.
    Answers are stored via the standard SurveyResponse/SurveyAnswer mechanism.
    One shared evaluation record per student (unique constraint).
    """
    try:
        ss = SpecialSurvey.query.get(special_survey_id)
        if not ss:
            return {'status': False, 'message': 'Umfrage nicht gefunden.'}
        if ss.status not in ('active',):
            return {'status': False, 'message': 'Die Umfrage ist nicht aktiv.'}
        if not ss.teacher_survey_id:
            return {'status': False, 'message': 'Kein Fragebogen-Template zugewiesen.'}
        if not survey_answers:
            return {'status': False, 'message': 'Keine Antworten übermittelt.'}

        # Verify teacher is assigned to this student's class
        student = SpecialSurveyStudent.query.get(student_id)
        if not student or student.special_survey_id != special_survey_id:
            return {'status': False, 'message': 'Schüler nicht gefunden.'}

        assignment = SpecialSurveyClassTeacher.query.filter_by(
            special_survey_id=special_survey_id,
            class_name=student.class_name,
            teacher_uuid=user_uuid,
        ).first()
        if not assignment:
            return {'status': False, 'message': 'Sie sind dieser Klasse nicht zugewiesen.'}

        # Find or create evaluation
        existing = SpecialSurveyTeacherEvaluation.query.filter_by(
            special_survey_id=special_survey_id,
            student_id=student_id,
        ).first()

        if existing:
            existing.teacher_uuid = user_uuid
        else:
            existing = SpecialSurveyTeacherEvaluation(
                special_survey_id=special_survey_id,
                student_id=student_id,
                teacher_uuid=user_uuid,
            )
            db.session.add(existing)
            db.session.flush()  # get id for the evaluation

        # Store answers via standard survey response mechanism
        _store_teacher_survey_response(ss, existing, user_uuid, student_id, survey_answers)

        db.session.commit()
        return {'status': True, 'message': 'Bewertung gespeichert.'}

    except Exception as e:
        db.session.rollback()
        print(f'[Surveys] Error: {e}')
        return {'status': False, 'message': 'Ein Fehler ist aufgetreten.'}


def _store_teacher_survey_response(ss, evaluation, user_uuid, student_id, survey_answers):
    """Store teacher evaluation answers via the standard survey response mechanism."""
    # Remove old response if re-submitting
    if evaluation.survey_response_id:
        old_resp = SurveyResponse.query.get(evaluation.survey_response_id)
        if old_resp:
            db.session.delete(old_resp)
            db.session.flush()

    response = SurveyResponse(
        survey_id=ss.teacher_survey_id,
        user_uuid=user_uuid,
    )
    db.session.add(response)
    db.session.flush()

    for a_data in survey_answers:
        answer = SurveyAnswer(
            response_id=response.id,
            question_id=a_data['question_id'],
            answer_text=a_data.get('answer_text'),
            selected_option_id=a_data.get('selected_option_id'),
            selected_option_ids=a_data.get('selected_option_ids'),
        )
        db.session.add(answer)

    evaluation.survey_response_id = response.id


# ── Excel Export ─────────────────────────────────────────────────────

def export_special_survey_xlsx(special_survey_id):
    """
    Generate an Excel file with one row per student containing all collected data.
    Teacher evaluation columns come dynamically from the linked survey template.
    Per-question Excel configuration (excel_config_json) controls cell formatting:
      - color_marker: highlight the student name cell with a color based on the answer
      - option_text: insert the selected option text (default behavior)
      - custom_text_mapping: insert a custom text string based on the answer
    Returns (BytesIO_buffer, filename) or (None, error_message).
    """
    import json
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    ss = SpecialSurvey.query.get(special_survey_id)
    if not ss:
        return None, 'Umfrage nicht gefunden.'

    wb = Workbook()
    ws = wb.active
    ws.title = 'Klassenzusammensetzung'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')

    # Base headers (student + wish data)
    base_headers = [
        'Schüler/in', 'Klasse', 'Wunsch 1', 'Wunsch 2',
        'Gewählter Elternaccount', 'Eltern bestätigt',
    ]

    # Dynamic teacher evaluation headers from template questions
    template_questions = []
    if ss.teacher_survey_id:
        survey = Survey.query.get(ss.teacher_survey_id)
        if survey:
            template_questions = sorted(survey.questions, key=lambda q: q.order)

    # Parse excel_config_json per question
    question_excel_configs = {}
    for q in template_questions:
        try:
            cfg = json.loads(q.excel_config_json or '{}')
        except (json.JSONDecodeError, TypeError):
            cfg = {}
        question_excel_configs[q.id] = cfg

    question_headers = [q.text for q in template_questions]
    headers = base_headers + question_headers

    ws.append(headers)
    for i, cell in enumerate(ws[1]):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Build lookup maps
    student_map = {s.id: s for s in ss.students}
    wish_map = {}
    for w in ss.wishes:
        wish_map[w.student_id] = w
    eval_map = {}
    for e in ss.evaluations:
        eval_map[e.student_id] = e

    # Build option lookup for choice questions
    option_map = {}
    for q in template_questions:
        for opt in q.options:
            option_map[opt.id] = opt.text

    def _resolve_answer_text(q, a):
        """Resolve a question's answer to its display text."""
        if not a:
            return ''
        if q.question_type == 'text':
            return a.answer_text or ''
        elif q.question_type == 'single_choice':
            return option_map.get(a.selected_option_id, '') if a.selected_option_id else ''
        elif q.question_type == 'multiple_choice':
            if a.selected_option_ids:
                ids = [int(x) for x in a.selected_option_ids.split(',') if x.strip()]
                return ', '.join(option_map.get(oid, '') for oid in ids)
            return ''
        elif q.question_type == 'rating':
            return a.answer_text or ''
        elif q.question_type == 'yes_no':
            return a.answer_text or ''
        return a.answer_text or ''

    def _apply_excel_config(cell, name_cell, answer_text, cfg):
        """
        Apply Excel export configuration to a cell.
        cfg structure:
          excel_output_type: 'color_marker' | 'option_text' | 'custom_text_mapping'
          color_mappings: { answer_text: '#RRGGBB', ... }
          text_mappings: { answer_text: 'Custom text', ... }
        """
        output_type = cfg.get('excel_output_type', 'option_text')

        if output_type == 'color_marker':
            # Color the name cell based on the answer
            color_mappings = cfg.get('color_mappings', {})
            color_hex = color_mappings.get(answer_text)
            if color_hex:
                # Strip '#' prefix if present
                color_hex = color_hex.lstrip('#')
                try:
                    fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
                    name_cell.fill = fill
                except Exception:
                    pass  # Skip invalid color codes
            # Still put the answer text in the question cell
            cell.value = answer_text

        elif output_type == 'custom_text_mapping':
            # Replace answer text with a custom mapped value
            text_mappings = cfg.get('text_mappings', {})
            mapped_text = text_mappings.get(answer_text, answer_text)
            cell.value = mapped_text

        else:  # 'option_text' or default
            cell.value = answer_text

    # One row per student, sorted by class then name
    sorted_students = sorted(ss.students, key=lambda s: (s.class_name, s.last_name, s.first_name))
    name_col_idx = 1  # Column A is the student name (1-indexed)

    for student in sorted_students:
        wish = wish_map.get(student.id)
        ev = eval_map.get(student.id)

        wish1_name = ''
        wish2_name = ''
        parent_name = ''
        parent_confirmed = ''

        if wish:
            if wish.wish1_student_id and wish.wish1_student_id in student_map:
                w1 = student_map[wish.wish1_student_id]
                wish1_name = f"{w1.first_name} {w1.last_name}"
            if wish.wish2_student_id and wish.wish2_student_id in student_map:
                w2 = student_map[wish.wish2_student_id]
                wish2_name = f"{w2.first_name} {w2.last_name}"
            if wish.selected_parent:
                parent_name = f"{wish.selected_parent.first_name} {wish.selected_parent.last_name} ({wish.selected_parent.account})"
            parent_confirmed = 'Ja' if wish.parent_confirmed else 'Nein'

        row = [
            f"{student.first_name} {student.last_name}",
            student.class_name,
            wish1_name,
            wish2_name,
            parent_name,
            parent_confirmed,
        ]

        # Add dynamic template question answers (placeholder, filled below with config)
        if ev and ev.survey_response_id and ev.survey_response:
            answer_by_qid = {a.question_id: a for a in ev.survey_response.answers}
            for q in template_questions:
                a = answer_by_qid.get(q.id)
                answer_text = _resolve_answer_text(q, a)
                row.append(answer_text)  # placeholder value
        else:
            row.extend([''] * len(template_questions))

        ws.append(row)
        current_row = ws.max_row
        name_cell = ws.cell(row=current_row, column=name_col_idx)

        # Now apply per-question Excel configurations
        if ev and ev.survey_response_id and ev.survey_response:
            answer_by_qid = {a.question_id: a for a in ev.survey_response.answers}
            for q_offset, q in enumerate(template_questions):
                cfg = question_excel_configs.get(q.id, {})
                if not cfg or not cfg.get('excel_output_type'):
                    continue  # No config, keep the default text
                a = answer_by_qid.get(q.id)
                answer_text = _resolve_answer_text(q, a)
                col_idx = len(base_headers) + q_offset + 1  # 1-indexed
                question_cell = ws.cell(row=current_row, column=col_idx)
                _apply_excel_config(question_cell, name_cell, answer_text, cfg)

    # Column widths
    base_widths = [25, 12, 25, 25, 30, 15]
    question_widths = [30] * len(template_questions)
    col_widths = base_widths + question_widths
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Overview sheet
    ws_overview = wb.create_sheet(title='Übersicht', index=0)
    ws_overview.append(['Spezialumfrage', ss.title])
    ws_overview.append(['Jahrgang', ss.grade_level])
    ws_overview.append(['Status', ss.status])
    ws_overview.append(['Schüler gesamt', len(ss.students)])
    ws_overview.append(['Elternaccounts', len(ss.parents)])
    ws_overview.append(['Klassenlehrkräfte', len(ss.class_teachers)])
    ws_overview.append([])

    # Phase summary
    total_wishes = len(ss.wishes)
    confirmed_wishes = sum(1 for w in ss.wishes if w.parent_confirmed)
    total_evals = len(ss.evaluations)
    ws_overview.append(['Abgegebene Wünsche', total_wishes])
    ws_overview.append(['Bestätigte Wünsche', confirmed_wishes])
    ws_overview.append(['Lehrerbewertungen', total_evals])

    ws_overview.column_dimensions['A'].width = 30
    ws_overview.column_dimensions['B'].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"spezialumfrage_{special_survey_id}_ergebnisse.xlsx"
    return buf, filename


# ── Lazy-link: match unlinked CSV records to logged-in users ────────

def _lazy_link_user(user):
    """
    If any SpecialSurveyStudent or SpecialSurveyParent rows have
    account == user.username but user_uuid IS NULL, link them now.
    This handles the common case where CSV import happened before the
    user's first OAuth login (which is when the User row is created).
    """
    changed = False
    unlinked_students = SpecialSurveyStudent.query.filter_by(
        account=user.username, user_uuid=None
    ).all()
    for s in unlinked_students:
        s.user_uuid = user.uuid
        changed = True

    unlinked_parents = SpecialSurveyParent.query.filter_by(
        account=user.username, user_uuid=None
    ).all()
    for p in unlinked_parents:
        p.user_uuid = user.uuid
        changed = True

    if changed:
        db.session.commit()


# ── Participant Management (list / add / remove) ────────────────────

def get_participants(special_survey_id, creator_uuid):
    """
    Return all participants (students + parents) with their survey data.
    Only the survey creator (or admin) can access this.
    """
    ss = SpecialSurvey.query.get(special_survey_id)
    if not ss:
        return None, 'Umfrage nicht gefunden.'
    if ss.creator_uuid != creator_uuid:
        return None, 'Keine Berechtigung.'

    # Build wish map and eval map
    wish_map = {}
    for w in ss.wishes:
        wish_map[w.student_id] = w.to_dict()
    eval_map = {}
    for e in ss.evaluations:
        eval_map[e.student_id] = {
            'teacher_name': e.teacher.username if e.teacher else None,
            'evaluated': True,
        }

    participants = []

    # Students
    for s in ss.students:
        wish = wish_map.get(s.id)
        ev = eval_map.get(s.id)
        participants.append({
            'participant_id': s.id,
            'role': 'student',
            'account': s.account,
            'first_name': s.first_name,
            'last_name': s.last_name,
            'display_name': f"{s.first_name} {s.last_name}",
            'class_name': s.class_name,
            'user_uuid': s.user_uuid,
            'linked': s.user_uuid is not None,
            'wish': wish,
            'evaluation': ev,
        })

    # Parents
    for p in ss.parents:
        # Check if any student selected this parent and whether they confirmed
        parent_wishes = SpecialSurveyStudentWish.query.filter_by(
            special_survey_id=special_survey_id,
            selected_parent_id=p.id,
        ).all()
        confirmed = any(w.parent_confirmed for w in parent_wishes)
        selected_by = []
        for w in parent_wishes:
            student = SpecialSurveyStudent.query.get(w.student_id)
            if student:
                selected_by.append(f"{student.first_name} {student.last_name}")

        participants.append({
            'participant_id': p.id,
            'role': 'parent',
            'account': p.account,
            'first_name': p.first_name,
            'last_name': p.last_name,
            'display_name': f"{p.first_name} {p.last_name}",
            'class_name': None,
            'email': p.email,
            'user_uuid': p.user_uuid,
            'linked': p.user_uuid is not None,
            'confirmed': confirmed,
            'selected_by': selected_by,
        })

    return participants, None


def remove_participant(special_survey_id, participant_id, role, creator_uuid):
    """
    Remove a student or parent from the special survey.
    Also cleans up any related wishes / evaluations.
    """
    try:
        ss, err = _get_survey_for_owner(special_survey_id, creator_uuid)
        if err:
            return err

        if role == 'student':
            student = SpecialSurveyStudent.query.get(participant_id)
            if not student or student.special_survey_id != special_survey_id:
                return {'status': False, 'message': 'Schüler/in nicht gefunden.'}

            # Delete related wishes (where student is the wisher)
            SpecialSurveyStudentWish.query.filter_by(
                special_survey_id=special_survey_id, student_id=student.id
            ).delete()
            # Null out wish references where this student was wished for
            affected_wishes = SpecialSurveyStudentWish.query.filter(
                SpecialSurveyStudentWish.special_survey_id == special_survey_id,
                (SpecialSurveyStudentWish.wish1_student_id == student.id) |
                (SpecialSurveyStudentWish.wish2_student_id == student.id)
            ).all()
            for w in affected_wishes:
                if w.wish1_student_id == student.id:
                    w.wish1_student_id = None
                if w.wish2_student_id == student.id:
                    w.wish2_student_id = None
            # Delete evaluations
            SpecialSurveyTeacherEvaluation.query.filter_by(
                special_survey_id=special_survey_id, student_id=student.id
            ).delete()

            name = f"{student.first_name} {student.last_name}"
            db.session.delete(student)
            db.session.commit()
            return {'status': True, 'message': f'Schüler/in "{name}" entfernt.'}

        elif role == 'parent':
            parent = SpecialSurveyParent.query.get(participant_id)
            if not parent or parent.special_survey_id != special_survey_id:
                return {'status': False, 'message': 'Elternteil nicht gefunden.'}

            # Clear parent selection from wishes
            wishes = SpecialSurveyStudentWish.query.filter_by(
                special_survey_id=special_survey_id, selected_parent_id=parent.id
            ).all()
            for w in wishes:
                w.selected_parent_id = None
                w.parent_confirmed = False
                w.locked = False

            name = f"{parent.first_name} {parent.last_name}"
            db.session.delete(parent)
            db.session.commit()
            return {'status': True, 'message': f'Elternteil "{name}" entfernt.'}

        else:
            return {'status': False, 'message': f'Unbekannte Rolle: {role}'}

    except Exception as e:
        db.session.rollback()
        print(f'[Surveys] Error: {e}')
        return {'status': False, 'message': 'Ein Fehler ist aufgetreten.'}


def add_participant(special_survey_id, creator_uuid, username, role, class_name=None):
    """
    Add a single participant by username and role.
    Looks up the user in the user table; stores account even if not yet linked.
    Only allowed when the survey is in setup or active state.
    """
    try:
        ss, err = _get_survey_for_owner(special_survey_id, creator_uuid)
        if err:
            return err
        if ss.status in ('completed', 'archived'):
            return {'status': False, 'message': 'Teilnehmer k\u00f6nnen in abgeschlossenen/archivierten Umfragen nicht hinzugef\u00fcgt werden.'}

        if not username or not username.strip():
            return {'status': False, 'message': 'Benutzername ist erforderlich.'}
        username = username.strip()

        # Check user exists in database
        user = User.query.filter_by(username=username).first()

        # Safe name extraction: handle usernames with or without dots
        def _parse_name(uname):
            parts = uname.split('.')
            if len(parts) >= 2:
                return parts[0].capitalize(), parts[-1].capitalize()
            return uname.capitalize(), uname.capitalize()

        if role == 'student':
            if not class_name or not class_name.strip():
                return {'status': False, 'message': 'Klasse ist erforderlich für Schüler.'}
            class_name = class_name.strip()

            # Check duplicate
            existing = SpecialSurveyStudent.query.filter_by(
                special_survey_id=special_survey_id, account=username
            ).first()
            if existing:
                return {'status': False, 'message': f'"{username}" ist bereits als Schüler/in eingetragen.'}

            first_name, last_name = _parse_name(user.username if user else username)
            student = SpecialSurveyStudent(
                special_survey_id=ss.id,
                account=username,
                first_name=first_name,
                last_name=last_name,
                class_name=class_name,
                user_uuid=user.uuid if user else None,
            )
            db.session.add(student)
            db.session.commit()
            return {'status': True, 'message': f'Schüler/in "{username}" hinzugefügt.', 'participant': student.to_dict()}

        elif role == 'parent':
            existing = SpecialSurveyParent.query.filter_by(
                special_survey_id=special_survey_id, account=username
            ).first()
            if existing:
                return {'status': False, 'message': f'"{username}" ist bereits als Elternteil eingetragen.'}

            first_name, last_name = _parse_name(user.username if user else username)
            parent = SpecialSurveyParent(
                special_survey_id=ss.id,
                account=username,
                first_name=first_name,
                last_name=last_name,
                email='',
                user_uuid=user.uuid if user else None,
            )
            db.session.add(parent)
            db.session.commit()
            return {'status': True, 'message': f'Elternteil "{username}" hinzugefügt.', 'participant': parent.to_dict()}

        else:
            return {'status': False, 'message': f'Unbekannte Rolle: {role}'}

    except Exception as e:
        db.session.rollback()
        print(f'[Surveys] Error: {e}')
        return {'status': False, 'message': 'Ein Fehler ist aufgetreten.'}


# ── Helper: Get active special surveys for a user ───────────────────

def get_active_special_surveys_for_user(user_uuid):
    """
    Return special surveys the user can participate in
    (as student in phase1, parent in phase2, or teacher in phase3).
    """
    user = User.query.filter_by(uuid=user_uuid).first()
    if not user:
        return []

    # Lazy-link: resolve any CSV-imported records whose user hadn't logged in yet
    _lazy_link_user(user)

    surveys = []

    # Student role: active surveys where user is a student
    student_records = SpecialSurveyStudent.query.filter_by(user_uuid=user_uuid).all()
    for sr in student_records:
        ss = sr.special_survey
        if ss.status == 'active' and not ss.is_deleted:
            wish = SpecialSurveyStudentWish.query.filter_by(
                special_survey_id=ss.id, student_id=sr.id
            ).first()
            surveys.append({
                'id': ss.id,
                'title': ss.title,
                'type': 'special',
                'phase': 1,
                'role': 'student',
                'description': 'Wählen Sie zwei Mitschüler/innen und ein Elternteil',
                'already_responded': wish is not None and wish.wish1_student_id is not None,
                'locked': wish.locked if wish else False,
            })

    # Parent role: active surveys where user is a parent AND was selected by a student
    parent_records = SpecialSurveyParent.query.filter_by(user_uuid=user_uuid).all()
    for pr in parent_records:
        ss = pr.special_survey
        if ss.status == 'active' and not ss.is_deleted:
            # Parent only sees surveys where they were selected by a student
            wishes = SpecialSurveyStudentWish.query.filter_by(
                special_survey_id=ss.id, selected_parent_id=pr.id
            ).all()
            if wishes:
                all_confirmed = all(w.parent_confirmed for w in wishes)
                surveys.append({
                    'id': ss.id,
                    'title': ss.title,
                    'type': 'special',
                    'phase': 2,
                    'role': 'parent',
                    'description': 'Bestätigen Sie die Wünsche Ihres Kindes',
                    'already_responded': all_confirmed,
                    'locked': False,
                })

    # Teacher role: active surveys where user is assigned as class teacher
    teacher_records = SpecialSurveyClassTeacher.query.filter_by(teacher_uuid=user_uuid).all()
    for tr in teacher_records:
        ss = tr.special_survey
        if ss.status == 'active' and not ss.is_deleted:
            students = SpecialSurveyStudent.query.filter_by(
                special_survey_id=ss.id, class_name=tr.class_name
            ).all()
            evaluated = SpecialSurveyTeacherEvaluation.query.filter(
                SpecialSurveyTeacherEvaluation.special_survey_id == ss.id,
                SpecialSurveyTeacherEvaluation.student_id.in_([s.id for s in students]),
            ).count()

            surveys.append({
                'id': ss.id,
                'title': ss.title,
                'type': 'special',
                'phase': 3,
                'role': 'teacher',
                'description': f'Bewerten Sie die Schüler/innen der Klasse {tr.class_name}',
                'already_responded': evaluated == len(students) and len(students) > 0,
                'locked': False,
                'progress': f'{evaluated}/{len(students)}',
            })

    # Deduplicate by survey id + role
    seen = set()
    deduped = []
    for s in surveys:
        key = (s['id'], s['role'])
        if key not in seen:
            seen.add(key)
            deduped.append(s)

    return deduped


# ── Add Participants Post-Creation ──────────────────────────────────

def add_students(special_survey_id, creator_uuid, csv_content):
    """Add additional students to an existing special survey from CSV."""
    try:
        ss, err = _get_survey_for_owner(special_survey_id, creator_uuid)
        if err:
            return err

        students_data, err = _parse_csv(csv_content, is_student=True)
        if err:
            return {'status': False, 'message': f'Schüler-CSV: {err}'}

        # Deduplicate: skip accounts already in this survey
        existing_accounts = {s.account for s in ss.students}
        added = 0
        skipped = 0

        for s_data in students_data:
            if s_data['account'] in existing_accounts:
                skipped += 1
                continue

            user = User.query.filter_by(username=s_data['account']).first()
            student = SpecialSurveyStudent(
                special_survey_id=ss.id,
                account=s_data['account'],
                first_name=s_data['first_name'],
                last_name=s_data['last_name'],
                class_name=s_data['class_name'],
                user_uuid=user.uuid if user else None,
            )
            db.session.add(student)
            existing_accounts.add(s_data['account'])
            added += 1

        db.session.commit()
        msg = f'{added} Schüler/innen hinzugefügt.'
        if skipped:
            msg += f' {skipped} bereits vorhanden (übersprungen).'
        return {'status': True, 'message': msg, 'added': added, 'skipped': skipped}

    except Exception as e:
        db.session.rollback()
        print(f'[Surveys] Error: {e}')
        return {'status': False, 'message': 'Ein Fehler ist aufgetreten.'}


def add_parents(special_survey_id, creator_uuid, csv_content):
    """Add additional parents to an existing special survey from CSV."""
    try:
        ss, err = _get_survey_for_owner(special_survey_id, creator_uuid)
        if err:
            return err

        parents_data, err = _parse_csv(csv_content, is_student=False)
        if err:
            return {'status': False, 'message': f'Eltern-CSV: {err}'}

        existing_accounts = {p.account for p in ss.parents}
        added = 0
        skipped = 0

        for p_data in parents_data:
            if p_data['account'] in existing_accounts:
                skipped += 1
                continue

            user = User.query.filter_by(username=p_data['account']).first()
            parent = SpecialSurveyParent(
                special_survey_id=ss.id,
                account=p_data['account'],
                first_name=p_data['first_name'],
                last_name=p_data['last_name'],
                email=p_data.get('email', ''),
                user_uuid=user.uuid if user else None,
            )
            db.session.add(parent)
            existing_accounts.add(p_data['account'])
            added += 1

        db.session.commit()
        msg = f'{added} Elternaccounts hinzugefügt.'
        if skipped:
            msg += f' {skipped} bereits vorhanden (übersprungen).'
        return {'status': True, 'message': msg, 'added': added, 'skipped': skipped}

    except Exception as e:
        db.session.rollback()
        print(f'[Surveys] Error: {e}')
        return {'status': False, 'message': 'Ein Fehler ist aufgetreten.'}


# ── Reset Student Wishes (Admin Override) ───────────────────────────

def reset_student_wishes(special_survey_id, student_id, creator_uuid):
    """
    Reset a student's wishes so they can re-submit.
    Removes the wish record (including parent confirmation and lock).
    Only the survey creator can do this while the survey is active.
    """
    try:
        ss, err = _get_survey_for_owner(special_survey_id, creator_uuid)
        if err:
            return err
        if ss.status in ('completed', 'archived'):
            return {'status': False, 'message': 'W\u00fcnsche k\u00f6nnen in abgeschlossenen/archivierten Umfragen nicht zur\u00fcckgesetzt werden.'}

        student = SpecialSurveyStudent.query.get(student_id)
        if not student or student.special_survey_id != special_survey_id:
            return {'status': False, 'message': 'Schüler/in nicht gefunden.'}

        wish = SpecialSurveyStudentWish.query.filter_by(
            special_survey_id=special_survey_id, student_id=student_id
        ).first()

        if not wish:
            return {'status': False, 'message': 'Keine Wünsche vorhanden zum Zurücksetzen.'}

        db.session.delete(wish)
        db.session.commit()

        return {
            'status': True,
            'message': f'Wünsche von {student.first_name} {student.last_name} zurückgesetzt.',
        }

    except Exception as e:
        db.session.rollback()
        print(f'[Surveys] Error: {e}')
        return {'status': False, 'message': 'Ein Fehler ist aufgetreten.'}


# ── Parent Notification Email ───────────────────────────────────────

def send_parent_notification_email(parent_id, special_survey_id, student_name, app_base_url=None):
    """
    Send an email notification to a parent when they are selected by a student.
    Uses SMTP settings from environment variables.
    Returns True on success, False on error (non-fatal).
    """
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    import os

    SMTP_TIMEOUT = 15  # seconds

    try:
        parent = SpecialSurveyParent.query.get(parent_id)
        if not parent or not parent.email:
            print(f"[Special Survey] Skipping email: parent {parent_id} has no email address.")
            return False

        ss = SpecialSurvey.query.get(special_survey_id)
        if not ss:
            print(f"[Special Survey] Skipping email: survey {special_survey_id} not found.")
            return False

        smtp_host = os.getenv('SMTP_HOST')
        smtp_port = int(os.getenv('SMTP_PORT', 587))
        smtp_user = os.getenv('SMTP_USER', '')
        smtp_pass = os.getenv('SMTP_PASSWORD', '')
        smtp_from = os.getenv('SMTP_FROM', smtp_user)

        if not smtp_host:
            print("[Special Survey] SMTP not configured – skipping parent email notification.")
            return False

        base_url = app_base_url or os.getenv('APP_BASE_URL', 'https://dashboard.hub.mdg-hamburg.de')
        survey_link = f"{base_url}/surveys/special/{special_survey_id}/phase2"

        subject = f"Neue Klassenzusammensetzung – Bestätigung erforderlich"
        # Escape all user-supplied values to prevent HTML injection
        safe_parent = html_escape(f"{parent.first_name} {parent.last_name}")
        safe_student = html_escape(student_name)
        safe_title = html_escape(ss.title)
        safe_link = html_escape(survey_link)
        body_html = f"""
        <html><body>
        <p>Liebe/r {safe_parent},</p>
        <p>Ihr Kind <strong>{safe_student}</strong> hat Sie im Rahmen der Umfrage
        <strong>„{safe_title}"</strong> als bestätigenden Elternteil gewählt.</p>
        <p>Bitte melden Sie sich an und bestätigen Sie die Schülerwünsche:</p>
        <p><a href="{safe_link}">{safe_link}</a></p>
        <p>Mit freundlichen Grüßen<br/>Ihr Schulteam</p>
        </body></html>
        """

        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From'] = smtp_from
        msg['To'] = parent.email
        msg.attach(MIMEText(body_html, 'html', 'utf-8'))

        with smtplib.SMTP(smtp_host, smtp_port, timeout=SMTP_TIMEOUT) as server:
            server.ehlo()
            if smtp_port != 25:
                server.starttls()
            if smtp_user and smtp_pass:
                server.login(smtp_user, smtp_pass)
            server.sendmail(smtp_from, [parent.email], msg.as_string())

        print(f"[Special Survey] Parent notification email sent to {parent.email}")
        return True

    except smtplib.SMTPException as e:
        print(f"[Special Survey] SMTP error sending to {parent_id}: {e}")
        return False
    except OSError as e:
        print(f"[Special Survey] Network error sending email to {parent_id}: {e}")
        return False
    except Exception as e:
        print(f"[Special Survey] Unexpected email error for parent {parent_id}: {e}")
        return False
