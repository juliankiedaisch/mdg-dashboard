"""
Normal Survey Routes

All API routes specific to normal survey CRUD, templates, sharing,
results, and participation. Extracted from the monolithic surveys.py.

Usage (called from the Module class):
    register_normal_routes(blueprint, module_url, oauth, socketio_ref)
"""
import io
from flask import session, request, jsonify, send_file
from src.decorators import login_required, permission_required
from src.permissions import user_has_permission
from src.db_models import User, Group
from src.db import db
from modules.surveys.common.db_models import Survey, SurveyResponse
from modules.surveys.common.db_functions import _parse_iso_dt, _utcnow_naive
from modules.surveys.normal.db_functions import (
    create_survey, update_survey, delete_survey,
    add_question, update_question, delete_question,
    submit_response, get_survey_results,
    edit_survey_full, share_template, clone_from_template,
    save_as_template, grant_edit_response, revoke_edit_response,
)


def _can_manage_normal():
    """Return True if the user can manage normal surveys / templates."""
    return (
        user_has_permission("surveys.manage.all")
        or user_has_permission("surveys.normal.manage")
    )


def register_normal_routes(blueprint, module_url, oauth, socketio_ref):
    """Register all normal-survey-specific API routes on the given blueprint."""

    # ── Survey CRUD ──

    @blueprint.route(f"/api{module_url}", methods=["GET"])
    @login_required(oauth)
    @permission_required("surveys.manage.all", "surveys.normal.manage")
    def list_surveys():
        """
        List surveys for the current user.
        Query params:
          - tab: 'active' (default) | 'templates' | 'archived' | 'trash'
          - template_type: 'normal' (default) | 'teacher_evaluation' | 'all'
        """
        tab = request.args.get('tab', 'active')
        user_uuid = session['user_uuid']
        tpl_type_filter = request.args.get('template_type', 'normal')

        if tab == 'trash':
            # Trash: soft-deleted surveys owned by this user
            if not (user_has_permission("surveys.delete.permanently")
                    or user_has_permission("surveys.admin")):
                return jsonify({'error': 'Keine Berechtigung.'}), 403
            surveys = Survey.query.filter(
                Survey.creator_uuid == user_uuid,
                Survey.is_deleted == True,
            ).order_by(Survey.deleted_at.desc()).all()
            return jsonify({"surveys": [s.to_dict() for s in surveys]})

        if tab == 'templates':
            def _tpl_type_condition(query):
                if tpl_type_filter == 'all':
                    return query
                return query.filter(Survey.template_type == tpl_type_filter)

            own_q = Survey.query.filter_by(
                creator_uuid=user_uuid, is_template=True
            ).filter(Survey.is_deleted == False)
            own = _tpl_type_condition(own_q).order_by(Survey.created_at.desc()).all()

            user = User.query.filter_by(uuid=user_uuid).first()
            shared_user_q = Survey.query.filter(
                Survey.is_template == True,
                Survey.is_deleted == False,
                Survey.shared_with_users.any(User.uuid == user_uuid),
            )
            shared_user = _tpl_type_condition(shared_user_q).all() if user else []

            user_group_ids = {g.id for g in user.groups} if user else set()
            shared_group_q = Survey.query.filter(
                Survey.is_template == True,
                Survey.is_deleted == False,
                Survey.shared_with_groups.any(Group.id.in_(user_group_ids)),
            )
            shared_group = _tpl_type_condition(shared_group_q).all() if user_group_ids else []

            seen = set()
            surveys = []
            for s in own + shared_user + shared_group:
                if s.id not in seen:
                    seen.add(s.id)
                    d = s.to_dict(include_questions=True)
                    d['is_owner'] = s.creator_uuid == user_uuid
                    surveys.append(d)

            return jsonify({"surveys": surveys})

        elif tab == 'archived':
            surveys = Survey.query.filter(
                Survey.creator_uuid == user_uuid,
                Survey.is_template == False,
                Survey.is_deleted == False,
                Survey.status == 'archived',
            ).order_by(Survey.created_at.desc()).all()
            return jsonify({"surveys": [s.to_dict() for s in surveys]})

        else:  # 'active'
            surveys = Survey.query.filter(
                Survey.creator_uuid == user_uuid,
                Survey.is_template == False,
                Survey.is_deleted == False,
                Survey.status.in_(['active', 'closed', 'draft']),
            ).order_by(Survey.created_at.desc()).all()
            return jsonify({"surveys": [s.to_dict() for s in surveys]})

    @blueprint.route(f"/api{module_url}", methods=["POST"])
    @login_required(oauth)
    @permission_required("surveys.manage.all", "surveys.normal.manage")
    def create_new_survey():
        """Create a survey or template with questions and group assignments."""
        data = request.get_json()
        title = data.get('title', '').strip()
        if not title:
            return jsonify({'status': False, 'message': 'Titel ist erforderlich.'}), 400
        if len(title) > 255:
            return jsonify({'status': False, 'message': 'Titel darf maximal 255 Zeichen lang sein.'}), 400
        description = data.get('description', '')
        if len(description) > 5000:
            return jsonify({'status': False, 'message': 'Beschreibung darf maximal 5000 Zeichen lang sein.'}), 400

        is_template = data.get('is_template', False)
        template_type = data.get('template_type', 'normal')
        if template_type not in ('normal', 'teacher_evaluation'):
            return jsonify({'status': False, 'message': 'Ungültiger Vorlagentyp.'}), 400
        try:
            starts_at = _parse_iso_dt(data.get('starts_at'))
            ends_at = _parse_iso_dt(data.get('ends_at'))
        except (ValueError, TypeError):
            return jsonify({'status': False, 'message': 'Ungültiges Datumsformat.'}), 400

        result = create_survey(
            title=title,
            description=data.get('description', ''),
            creator_uuid=session['user_uuid'],
            anonymous=data.get('anonymous', False),
            starts_at=starts_at if not is_template else None,
            ends_at=ends_at if not is_template else None,
            group_ids=data.get('group_ids', []) if not is_template else [],
            questions=data.get('questions', []),
            is_template=is_template,
            allow_edit_response=data.get('allow_edit_response', False),
            template_type=template_type,
        )

        socketio_ref.emit('load_menu', namespace='/main')
        code = 201 if result['status'] else 400
        return jsonify(result), code

    @blueprint.route(f"/api{module_url}/<int:survey_id>", methods=["GET"])
    @login_required(oauth)
    @permission_required("surveys.manage.all", "surveys.normal.manage")
    def get_survey(survey_id):
        """Get a single survey with full question data."""
        survey = Survey.query.get_or_404(survey_id)
        if survey.creator_uuid != session['user_uuid'] \
                and not user_has_permission("surveys.admin"):
            return jsonify({'error': 'Keine Berechtigung.'}), 403
        return jsonify({"survey": survey.to_dict(include_questions=True)})

    @blueprint.route(f"/api{module_url}/<int:survey_id>", methods=["PUT"])
    @login_required(oauth)
    @permission_required("surveys.manage.all", "surveys.normal.manage")
    def update_existing_survey(survey_id):
        """Update survey metadata."""
        data = request.get_json()
        result = update_survey(survey_id, data, session['user_uuid'])
        socketio_ref.emit('load_menu', namespace='/main')
        code = 200 if result['status'] else 400
        return jsonify(result), code

    @blueprint.route(f"/api{module_url}/<int:survey_id>", methods=["DELETE"])
    @login_required(oauth)
    @permission_required("surveys.manage.all", "surveys.normal.manage")
    def delete_existing_survey(survey_id):
        """Delete a survey."""
        is_admin = user_has_permission("surveys.admin")
        result = delete_survey(survey_id, session['user_uuid'], is_admin)
        socketio_ref.emit('load_menu', namespace='/main')
        code = 200 if result['status'] else 400
        return jsonify(result), code

    # ── Questions ──

    @blueprint.route(f"/api{module_url}/<int:survey_id>/questions", methods=["POST"])
    @login_required(oauth)
    @permission_required("surveys.manage.all", "surveys.normal.manage")
    def add_survey_question(survey_id):
        """Add a question to a survey."""
        q_data = request.get_json()
        if not q_data.get('text', '').strip():
            return jsonify({'status': False, 'message': 'Fragetext ist erforderlich.'}), 400
        result = add_question(survey_id, q_data, session['user_uuid'])
        code = 201 if result['status'] else 400
        return jsonify(result), code

    @blueprint.route(f"/api{module_url}/questions/<int:question_id>", methods=["PUT"])
    @login_required(oauth)
    @permission_required("surveys.manage.all", "surveys.normal.manage")
    def update_survey_question(question_id):
        """Update a question."""
        q_data = request.get_json()
        result = update_question(question_id, q_data, session['user_uuid'])
        code = 200 if result['status'] else 400
        return jsonify(result), code

    @blueprint.route(f"/api{module_url}/questions/<int:question_id>", methods=["DELETE"])
    @login_required(oauth)
    @permission_required("surveys.manage.all", "surveys.normal.manage")
    def delete_survey_question(question_id):
        """Delete a question."""
        result = delete_question(question_id, session['user_uuid'])
        code = 200 if result['status'] else 400
        return jsonify(result), code

    # ── Participation ──

    @blueprint.route(f"/api{module_url}/<int:survey_id>/participate", methods=["GET"])
    @login_required(oauth)
    def get_survey_for_participation(survey_id):
        """Get a survey for a participant."""
        survey = Survey.query.get_or_404(survey_id)
        if survey.status != 'active':
            return jsonify({'error': 'Umfrage ist nicht aktiv.'}), 400

        now = _utcnow_naive()
        if survey.starts_at and survey.starts_at > now:
            return jsonify({'error': 'Die Umfrage hat noch nicht begonnen.'}), 400
        if survey.ends_at and survey.ends_at <= now:
            return jsonify({'error': 'Die Umfrage ist bereits abgelaufen.', 'is_expired': True}), 400

        user = User.query.filter_by(uuid=session['user_uuid']).first()
        user_group_ids = {g.id for g in user.groups} if user else set()
        survey_group_ids = {g.id for g in survey.groups}
        if survey_group_ids and not (survey_group_ids & user_group_ids):
            return jsonify({'error': 'Sie sind nicht berechtigt, an dieser Umfrage teilzunehmen.'}), 403

        visible_questions = []
        for q in survey.questions:
            q_group_ids = {g.id for g in q.groups}
            if not q_group_ids or q_group_ids & user_group_ids:
                visible_questions.append(q.to_dict())

        existing = SurveyResponse.query.filter_by(
            survey_id=survey_id, user_uuid=session['user_uuid']
        ).first()
        if existing:
            if survey.allow_edit_response or existing.edit_granted:
                existing_answers = {}
                for ans in existing.answers:
                    existing_answers[ans.question_id] = ans.to_dict()
                return jsonify({
                    "survey": {
                        "id": survey.id,
                        "title": survey.title,
                        "description": survey.description,
                        "anonymous": survey.anonymous,
                        "allow_edit_response": survey.allow_edit_response,
                    },
                    "questions": visible_questions,
                    "existing_answers": existing_answers,
                    "is_edit": True,
                })
            return jsonify({'error': 'Sie haben bereits teilgenommen.', 'already_responded': True}), 400

        return jsonify({
            "survey": {
                "id": survey.id,
                "title": survey.title,
                "description": survey.description,
                "anonymous": survey.anonymous,
                "allow_edit_response": survey.allow_edit_response,
            },
            "questions": visible_questions,
        })

    @blueprint.route(f"/api{module_url}/<int:survey_id>/respond", methods=["POST"])
    @login_required(oauth)
    def submit_survey_response(survey_id):
        """Submit answers for a survey."""
        survey = Survey.query.get_or_404(survey_id)
        if survey.status != 'active':
            return jsonify({'status': False, 'message': 'Umfrage ist nicht aktiv.'}), 400
        now = _utcnow_naive()
        if survey.starts_at and survey.starts_at > now:
            return jsonify({'status': False, 'message': 'Die Umfrage hat noch nicht begonnen.'}), 400
        if survey.ends_at and survey.ends_at <= now:
            return jsonify({'status': False, 'message': 'Die Umfrage ist bereits abgelaufen.'}), 400

        user = User.query.filter_by(uuid=session['user_uuid']).first()
        user_group_ids = {g.id for g in user.groups} if user else set()
        survey_group_ids = {g.id for g in survey.groups}
        if survey_group_ids and not (survey_group_ids & user_group_ids):
            return jsonify({'status': False, 'message': 'Sie sind nicht berechtigt, an dieser Umfrage teilzunehmen.'}), 403

        data = request.get_json()
        answers = data.get('answers', [])
        if not answers:
            return jsonify({'status': False, 'message': 'Keine Antworten gesendet.'}), 400

        result = submit_response(survey_id, session['user_uuid'], answers)
        code = 200 if result['status'] else 400
        return jsonify(result), code

    # ── Results / Evaluation ──

    @blueprint.route(f"/api{module_url}/<int:survey_id>/results", methods=["GET"])
    @login_required(oauth)
    @permission_required("surveys.manage.all", "surveys.normal.manage")
    def survey_results(survey_id):
        """Get aggregated results for a survey."""
        survey = Survey.query.get_or_404(survey_id)
        if survey.creator_uuid != session['user_uuid'] \
                and not user_has_permission("surveys.admin"):
            return jsonify({'error': 'Keine Berechtigung.'}), 403

        results = get_survey_results(survey_id)
        if results is None:
            return jsonify({'error': 'Umfrage nicht gefunden.'}), 404
        return jsonify({"results": results})

    @blueprint.route(f"/api{module_url}/<int:survey_id>/results/xlsx", methods=["GET"])
    @login_required(oauth)
    @permission_required("surveys.manage.all", "surveys.normal.manage")
    def survey_results_xlsx(survey_id):
        """Download survey results as XLSX."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        survey = Survey.query.get_or_404(survey_id)
        if survey.creator_uuid != session['user_uuid'] \
                and not user_has_permission("surveys.admin"):
            return jsonify({'error': 'Keine Berechtigung.'}), 403

        results = get_survey_results(survey_id)
        if results is None:
            return jsonify({'error': 'Umfrage nicht gefunden.'}), 404

        wb = Workbook()
        ws_overview = wb.active
        ws_overview.title = 'Übersicht'
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')

        ws_overview.append(['Umfrage', results['title']])
        ws_overview.append(['Antworten gesamt', results['response_count']])
        ws_overview.append(['Anonym', 'Ja' if results.get('anonymous') else 'Nein'])
        ws_overview.append([])

        if not results.get('anonymous') and results.get('participants'):
            ws_overview.append(['Teilnehmer', 'Eingereicht am'])
            for cell in ws_overview[ws_overview.max_row]:
                cell.font = header_font
                cell.fill = header_fill
            for p in results['participants']:
                ws_overview.append([p['username'], p.get('submitted_at', '')])

        ws_overview.column_dimensions['A'].width = 30
        ws_overview.column_dimensions['B'].width = 30

        ws_results = wb.create_sheet(title='Ergebnisse')
        for q in results['questions']:
            ws_results.append([q['text']])
            row = ws_results[ws_results.max_row]
            row[0].font = Font(bold=True, size=12)

            q_type = q['question_type']
            if q_type in ('single_choice', 'multiple_choice') and q.get('option_results'):
                ws_results.append(['Option', 'Anzahl'])
                for cell in ws_results[ws_results.max_row]:
                    cell.font = header_font
                    cell.fill = header_fill
                for opt in q['option_results']:
                    ws_results.append([opt['text'], opt['count']])
            elif q_type == 'rating':
                ws_results.append(['Durchschnitt', q.get('average', 0)])
                ws_results.append(['Anzahl Bewertungen', q.get('answers_count', 0)])
            elif q_type == 'yes_no':
                ws_results.append(['Ja', q.get('yes_count', 0)])
                ws_results.append(['Nein', q.get('no_count', 0)])
            elif q_type == 'text':
                ws_results.append(['Antworten'])
                ws_results[ws_results.max_row][0].font = header_font
                ws_results[ws_results.max_row][0].fill = header_fill
                for txt in q.get('text_answers', []):
                    ws_results.append([txt])

            if not results.get('anonymous') and q.get('user_answers'):
                ws_results.append([])
                ws_results.append(['Teilnehmer', 'Antwort'])
                for cell in ws_results[ws_results.max_row]:
                    cell.font = header_font
                    cell.fill = header_fill
                for ua in q['user_answers']:
                    ws_results.append([ua['username'], ua['answer']])

            ws_results.append([])

        ws_results.column_dimensions['A'].width = 35
        ws_results.column_dimensions['B'].width = 30

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"umfrage_{survey_id}_ergebnisse.xlsx"
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename,
        )

    # ── Response edit grants ──

    @blueprint.route(f"/api{module_url}/responses/<int:response_id>/grant-edit", methods=["PUT"])
    @login_required(oauth)
    @permission_required("surveys.manage.all", "surveys.normal.manage")
    def grant_response_edit(response_id):
        """Grant a one-time edit permission for a specific response."""
        result = grant_edit_response(response_id, session['user_uuid'])
        code = 200 if result['status'] else 400
        return jsonify(result), code

    @blueprint.route(f"/api{module_url}/responses/<int:response_id>/revoke-edit", methods=["PUT"])
    @login_required(oauth)
    @permission_required("surveys.manage.all", "surveys.normal.manage")
    def revoke_response_edit(response_id):
        """Revoke a previously granted one-time edit permission."""
        result = revoke_edit_response(response_id, session['user_uuid'])
        code = 200 if result['status'] else 400
        return jsonify(result), code

    # ── Full edit (non-active surveys) ──

    @blueprint.route(f"/api{module_url}/<int:survey_id>/edit", methods=["PUT"])
    @login_required(oauth)
    @permission_required("surveys.manage.all", "surveys.normal.manage")
    def full_edit_survey(survey_id):
        """Full edit of a survey (metadata + questions). Blocked when active."""
        data = request.get_json()
        if 'title' in data and len(data['title'].strip()) > 255:
            return jsonify({'status': False, 'message': 'Titel darf maximal 255 Zeichen lang sein.'}), 400
        if 'description' in data and len(data.get('description', '')) > 5000:
            return jsonify({'status': False, 'message': 'Beschreibung darf maximal 5000 Zeichen lang sein.'}), 400
        result = edit_survey_full(survey_id, data, session['user_uuid'])
        socketio_ref.emit('load_menu', namespace='/main')
        code = 200 if result['status'] else 400
        return jsonify(result), code

    # ── Template sharing ──

    @blueprint.route(f"/api{module_url}/<int:survey_id>/share", methods=["PUT"])
    @login_required(oauth)
    @permission_required("surveys.manage.all", "surveys.normal.manage")
    def share_survey_template(survey_id):
        """Share a template with groups and/or users."""
        data = request.get_json()
        result = share_template(
            survey_id, session['user_uuid'],
            group_ids=data.get('group_ids'),
            user_uuids=data.get('user_uuids'),
        )
        code = 200 if result['status'] else 400
        return jsonify(result), code

    # ── Clone from template ──

    @blueprint.route(f"/api{module_url}/from-template/<int:template_id>", methods=["POST"])
    @login_required(oauth)
    @permission_required("surveys.manage.all", "surveys.normal.manage")
    def create_from_template(template_id):
        """Create a new survey by cloning a template."""
        data = request.get_json() or {}
        try:
            starts_at = _parse_iso_dt(data.get('starts_at'))
            ends_at = _parse_iso_dt(data.get('ends_at'))
        except (ValueError, TypeError):
            return jsonify({'status': False, 'message': 'Ungültiges Datumsformat.'}), 400
        result = clone_from_template(
            template_id, session['user_uuid'],
            title=data.get('title'),
            group_ids=data.get('group_ids', []),
            starts_at=starts_at,
            ends_at=ends_at,
            anonymous=data.get('anonymous'),
        )
        socketio_ref.emit('load_menu', namespace='/main')
        code = 201 if result['status'] else 400
        return jsonify(result), code

    # ── Save as template ──

    @blueprint.route(f"/api{module_url}/<int:survey_id>/save-as-template", methods=["POST"])
    @login_required(oauth)
    @permission_required("surveys.manage.all", "surveys.normal.manage")
    def save_survey_as_template(survey_id):
        """Clone any survey into a new template."""
        result = save_as_template(survey_id, session['user_uuid'])
        code = 201 if result['status'] else 400
        return jsonify(result), code
